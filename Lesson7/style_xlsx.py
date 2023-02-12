import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
class Style:

    def styling(self):
        filename = 'PhoneBook.xlsx'
        book = openpyxl.load_workbook(filename)
        sheet = book.active

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 50

        list_col = ['A','B','C']

        font = Font(name='Calibri', size=12, bold=True, italic=True, color='00000000')
        fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid")

        for i in list_col:
            sheet[f'{i}1'].font = font
            sheet[f'{i}1'].fill = fill

        rows = sheet.max_row

        thin = Side(border_style="thin", color="000000")
        for row in sheet[f'A1:C{rows}']:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


        book.save('PhoneBook.xlsx')
